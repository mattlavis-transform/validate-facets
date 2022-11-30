import os
import json
import pyperclip
import sys
from openpyxl import load_workbook
from dotenv import load_dotenv


class FacetList:
    def __init__(self):
        load_dotenv('.env')
        self.facets_source = os.getenv('FACETS_SOURCE')
        self.classifier_folder = os.getenv('CLASSIFIER_FOLDER')

        self.load_data()

    def number_to_letter(self, number):
        number = number + 64
        return chr(number)

    def load_data(self):
        self.facet_addresses = {}
        wb = load_workbook(filename=self.facets_source)

        # Get the full list of facets in the facets sheet
        sht_facets = wb['facets']
        row_count = sht_facets.max_row
        master_facets = []
        for row in range(2, row_count):
            cell = sht_facets.cell(row=row, column=1)
            if cell.value is not None:
                master_facets.append(cell.value)

        master_facets.sort()

        sht_headings = wb['headings']
        row_count = sht_headings.max_row
        col_count = sht_headings.max_column

        error_count = 0
        facets = []
        for row in range(2, row_count):
            my_facets = []
            for column in range(5, col_count):
                cell = sht_headings.cell(row=row, column=column)
                if cell.value is not None:
                    value = cell.value.strip()
                    if len(value) > 2:
                        if value not in my_facets:
                            my_facets.append(value)
                            value = value.strip()
                            cell_address = self.number_to_letter(column) + str(row)
                            if value not in self.facet_addresses:
                                self.facet_addresses[value] = [cell_address]
                            else:
                                self.facet_addresses[value].append(cell_address)
                        else:
                            print("On row", str(row), "facet", value, "is duplicated")
                    value_lower = value.lower()
                    if value != "":
                        if " " in value or "/" in value or "-" in value or "%" in value:
                            error_count += 1
                        if value_lower != value:
                            error_count += 1

                        facets.append(value.strip())

        # Write all facets
        facet_output = os.path.join(os.getcwd(), "resources", "all_facets.txt")
        facets.sort()
        with open(facet_output, 'w') as f:
            for facet in facets:
                f.write(facet + "\n")
        f.close()

        # Write unique facets and write the classifier file if it is missing
        unique_facets = list(set(facets))
        unique_facets.sort()
        facet_output = os.path.join(os.getcwd(), "resources", "unique_facets.txt")
        with open(facet_output, 'w') as f:
            for facet in unique_facets:
                facet_filename = "classifier_" + facet + ".json"
                facet_filepath = os.path.join(self.classifier_folder, facet_filename)
                file_exists = os.path.exists(facet_filepath)
                if not file_exists:
                    rubric = [
                        {
                            "key": [
                                "trigger"
                            ]
                        }
                    ]
                    with open(facet_filepath, 'w') as f2:
                        json.dump(rubric, f2, indent=4)
                    f2.close()
                a = 1
                f.write(facet + "\n")
        f.close()

        # Write unique facets and addresses
        facet_output = os.path.join(os.getcwd(), "resources", "facet_addresses.txt")
        with open(facet_output, 'w') as f:
            # for facet in self.facet_addresses:
            for key in sorted(self.facet_addresses):
                facet = self.facet_addresses[key]
                f.write(key + ", " + str(len(self.facet_addresses[key])) + ", ")
                f.write(", ".join(self.facet_addresses[key]))
                f.write("\n")

        f.close()

        # Check that all of the noted facets on the headings table are also on the facets sheet
        unique_facets = list(set(facets))
        unique_facets.sort()
        missing_facet_output = os.path.join(os.getcwd(), "resources", "missing_facets.txt")
        with open(missing_facet_output, 'w') as f:
            for facet in unique_facets:
                if facet not in master_facets:
                    f.write(facet + "\n")
        f.close()

        print("\n\n", "Number of errors is", str(error_count), "\n\n")

        for row in range(2, row_count):
            for column in range(5, col_count):
                cell = sht_headings.cell(row=row, column=column)
                if cell.value is not None:
                    value = cell.value.strip()
                    value_lower = value.lower()
                    if value != "":
                        if " " in value or "/" in value or "-" in value or "%" in value:
                            print("Space or slash in value: ", value, " - on cell", str(row), ",", str(column))
                            pyperclip.copy(value)
                            sys.exit()
                        if value_lower != value:
                            print("Case mismatch: ", value, " - on cell", str(row), ",", str(column))
                            pyperclip.copy(value)
                            sys.exit()
